import openpyxl
import numpy as np
import pandas as pd


def sale_customer(path):
    wb = openpyxl.open(path)
    ws = wb.worksheets[0]
    info_sales_customer = {}
    for row in ws:
        if row[0].value in info_sales_customer.keys():
            continue
        else:
            info_sales_customer[row[0].value] = row[1].value
    ws = wb.create_sheet()
    for k in info_sales_customer:
        data_list = []
        data_list.append(k)
        data_list.append(info_sales_customer[k])
        ws.append(data_list)
    ws['c1'] = '2019及以前'
    ws['d1'] = '2021年'
    ws['e1'] = '2022年'
    ws['f1'] = '2023年'
    wb.close()
    wb.save(f'目标输出/{path}')

df = pd.read_excel(f'数据源.xlsx',sheet_name='Sheet1')

# print(df1[(df1['销售员'] == df2['销售员'])&(df1['客户名称'] == df2['客户名称'])]['实际应收'].sum())
df['应收产生年份'] = df['日期'].apply(lambda x:x.year)
# df[(df['应收超期天数']>180)]
# data_180 = df[(df['应收超期天数']>180)]
# cust = data_180.groupby(['客户名称','销售员','年',])['实际应收'].sum()


if __name__ == '__main__':
    pass
    #sale_customer('数据源.xlsx')